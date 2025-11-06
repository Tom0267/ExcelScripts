from openpyxl.utils import column_index_from_string
from pathHelper import ensure_user_csv
from openpyxl import load_workbook
from pathlib import Path
import sys
import os

def matcher(FILENAME):
    downloads_folder = str(Path.home() / "Downloads")
    output_dir = os.path.join(downloads_folder, "StationMatcher")
    os.makedirs(output_dir, exist_ok=True)

    ext = Path(FILENAME).suffix.lower()

    if ext not in [".xlsx", ".xlsm"]:
        raise ValueError(
            f"Unsupported input '{ext}'. Please save the file as .xlsx or .xlsm and run again."
        )

    keep_vba = (ext == ".xlsm")
    wb = load_workbook(FILENAME, data_only=False, keep_vba=keep_vba)

    #get stationMap from the stationMap.csv file
    station_map = {}
    csv_path = ensure_user_csv()
    with open(csv_path, "r", encoding="utf-8") as f:
        for idx, line in enumerate(f):
            if idx < 1:  # start from index 1 (skip header / first line)
                continue
            parts = line.strip().split(",")
            if len(parts) >= 2:
                try:
                    station_map[int(parts[0])] = parts[1]
                except ValueError:
                    continue

    def _to_int(val):
        if val is None:
            return None
        try:
            if isinstance(val, float):
                return int(val)
            return int(str(val).strip())
        except (ValueError, TypeError):
            return None

    for ws in wb.worksheets:
        # headers assumed on row 1
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        headers = {(c.value or ""): c.col_idx for c in header_cells}

        dep_hdr, arr_hdr = "RailDepartureCode", "RailArrivalCode"
        dep_col = headers.get(dep_hdr)
        arr_col = headers.get(arr_hdr)

        dest_dep_col = headers.get("AM") or column_index_from_string("AM")  # 39
        dest_arr_col = headers.get("AN") or column_index_from_string("AN")  # 40

        if not dep_col and not arr_col:
            continue

        for r in range(2, ws.max_row + 1):
            if dep_col:
                dep_key = _to_int(ws.cell(row=r, column=dep_col).value)
                if dep_key is not None and dep_key in station_map:
                    ws.cell(row=r, column=dest_dep_col, value=station_map[dep_key])
            if arr_col:
                arr_key = _to_int(ws.cell(row=r, column=arr_col).value)
                if arr_key is not None and arr_key in station_map:
                    ws.cell(row=r, column=dest_arr_col, value=station_map[arr_key])

    # Preserve the same extension as input (.xlsx stays .xlsx; .xlsm stays .xlsm)
    output_path = os.path.join(output_dir, Path(FILENAME).stem + "_matched" + ext)
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python match.py <FILENAME>")
    else:
        print(matcher(sys.argv[1]))